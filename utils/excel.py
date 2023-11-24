from typing import Generator, Any

import openpyxl
from fastapi import UploadFile, HTTPException
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from starlette import status

from schemas.user import IUserCreateByExcel


def _verify_upload_student_excel_format() -> bool:
    return True


def get_student_from_upload_excel(file: UploadFile) -> Generator[IUserCreateByExcel, None, None]:
    return _get_student_from_excel(file.file)


def _get_student_from_excel(file: Any) -> Generator[IUserCreateByExcel, None, None]:
    wb: Workbook = openpyxl.load_workbook(filename=file, read_only=True)

    sheet: Worksheet = wb[wb.sheetnames[0]]
    if not _verify_upload_student_excel_format():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="文件格式不正确")

    for row in sheet.iter_rows(min_col=1, min_row=2, values_only=True):
        if not row[0]:
            continue

        user = IUserCreateByExcel(
            account=str(row[0]),
            password=str(row[0]),
            name=row[1],
            gender=0 if row[2] == '女' else 1,
            class_name=row[3],
            duties=row[4]
        )
        yield user

    wb.close()


if __name__ == '__main__':
    for i in _get_student_from_excel('../files/班级名单空.xlsx'):
        print(i)
